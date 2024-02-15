


import random
import math
# import openpyxl
from openpyxl import load_workbook

class Sudoku:
    def __init__(self, N, K, difficulty, template_file):
        self.N = N
        self.K = K
        self.difficulty = difficulty
        self.template_file = template_file
        SRNd = math.sqrt(N)
        self.SRN = int(SRNd)
        self.mat = [[0 for _ in range(N)] for _ in range(N)]

    def fillValues(self):
        self.fillDiagonal()
        self.fillRemaining(0, self.SRN)
        self.removeKDigits()

    def fillDiagonal(self):
        for i in range(0, self.N, self.SRN):
            self.fillBox(i, i)

    def unUsedInBox(self, rowStart, colStart, num):
        for i in range(self.SRN):
            for j in range(self.SRN):
                if self.mat[rowStart + i][colStart + j] == num:
                    return False
        return True

    def fillBox(self, row, col):
        num = 0
        for i in range(self.SRN):
            for j in range(self.SRN):
                while True:
                    num = self.randomGenerator(self.N)
                    if self.unUsedInBox(row, col, num):
                        break
                self.mat[row + i][col + j] = num

    def randomGenerator(self, num):
        return math.floor(random.random() * num + 1)

    def checkIfSafe(self, i, j, num):
        return (
            self.unUsedInRow(i, num)
            and self.unUsedInCol(j, num)
            and self.unUsedInBox(i - i % self.SRN, j - j % self.SRN, num)
        )

    def unUsedInRow(self, i, num):
        for j in range(self.N):
            if self.mat[i][j] == num:
                return False
        return True

    def unUsedInCol(self, j, num):
        for i in range(self.N):
            if self.mat[i][j] == num:
                return False
        return True

    def fillRemaining(self, i, j):
        if i == self.N - 1 and j == self.N:
            return True

        if j == self.N:
            i += 1
            j = 0

        if self.mat[i][j] != 0:
            return self.fillRemaining(i, j + 1)

        for num in range(1, self.N + 1):
            if self.checkIfSafe(i, j, num):
                self.mat[i][j] = num
                if self.fillRemaining(i, j + 1):
                    return True
                self.mat[i][j] = 0

        return False

    def removeKDigits(self):
        # Adjust the number of digits to be removed based on difficulty
        if self.difficulty == 1:
            self.K += 0
        elif self.difficulty == 2:
            self.K += 10
        elif self.difficulty == 3:
            self.K += 15
        elif self.difficulty == 4:
            self.K += 20

        count = self.K
        while count != 0:
            i = self.randomGenerator(self.N) - 1
            j = self.randomGenerator(self.N) - 1
            if self.mat[i][j] != 0:
                count -= 1
                self.mat[i][j] = 0

    def solveSudoku(self):
        def findEmptyLocation():
            for i in range(self.N):
                for j in range(self.N):
                    if self.mat[i][j] == 0:
                        return i, j
            return -1, -1

        def isSafe(row, col, num):
            return (
                self.unUsedInRow(row, num)
                and self.unUsedInCol(col, num)
                and self.unUsedInBox(row - row % self.SRN, col - col % self.SRN, num)
            )

        def solve():
            i, j = findEmptyLocation()
            if i == -1:
                return True

            for num in range(1, self.N + 1):
                if isSafe(i, j, num):
                    self.mat[i][j] = num
                    if solve():
                        return True
                    self.mat[i][j] = 0

            return False

        solve()

    def printSudoku(self):
        for i in range(self.N):
            for j in range(self.N):
                print(self.mat[i][j], end=" ")
            print()

    def write_to_excel(self, excel_file):
        template_wb = load_workbook(self.template_file)
        ws = template_wb.active

        # Write generated Sudoku to Excel
        for i in range(self.N):
            for j in range(self.N):
                value = "" if self.mat[i][j] == 0 else self.mat[i][j]
                ws.cell(row=i + 1, column=j + 1, value=value)

        # Write solution to Excel
        self.solveSudoku()
        for i in range(self.N):
            for j in range(self.N):
                ws.cell(row=i + 1, column=j + self.N + 2, value=self.mat[i][j])

        template_wb.save(excel_file)
        print(f"\nSudoku and Solution written to {excel_file}")

def generate_sudoku_and_save():
    try:
        num_sudoku = int(input("Enter the number of Sudoku puzzles to generate: "))
        if num_sudoku <= 0:
            print("Please enter a positive integer.")
            return

        difficulty = int(input("Enter the difficulty level (1 to 4): "))
        if difficulty < 1 or difficulty > 4:
            print("Difficulty level must be between 1 and 4.")
            return

        template_file = "sample.xlsx"

        for i in range(1, num_sudoku + 1):
            excel_file = f"sudoku_solution_{i}_level_{difficulty}.xlsx"
            sudoku = Sudoku(N=9, K=40, difficulty=difficulty, template_file=template_file)
            sudoku.fillValues()
            sudoku.write_to_excel(excel_file)

        print(f"\n{num_sudoku} Sudoku puzzles and solutions written to individual Excel files using the template.")

    except ValueError:
        print("Error: Please enter valid integers.")

if __name__ == "__main__":
    generate_sudoku_and_save()
